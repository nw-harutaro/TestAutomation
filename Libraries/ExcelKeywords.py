from robot.api.deco import library, keyword
from openpyxl import load_workbook
import General
import os
import pandas as pd

@library(scope="SUITE")
class ExcelKeywords():

    def __init__(self):
        self.connection = None
    
    @keyword
    def get_last_row_number(self, file_path, sheet_name, row_number, col_number):
        
        try:
            workbook = load_workbook(filename=file_path, data_only=True)
            sheet = workbook[sheet_name]

            row_number = int(row_number)
            col_number = int(col_number) 

            # 指定された列の最後の行番号を取得
            last_row = 0
            for row in range(row_number, sheet.max_row + 1):
                if sheet.cell(row, col_number).value is not None:
                    last_row = row

            return last_row
        
        except Exception as e:
            # エラーが発生した場合に警告のポップアップを表示            
            General.show_warning_popup(f"エラーが発生しました: {str(e)}")
        

    @keyword
    def load_device_info_from_csv(self, file_path, log_folder, timestamp):        

        try:
            df = pd.read_csv(file_path, encoding="utf-8")               
            
            log_dir = os.path.join("Output", log_folder)
            log_dir = os.path.join(log_dir, timestamp)
            os.makedirs(log_dir, exist_ok=True)  # フォルダが存在しない場合は作成            

            DeviceInfo = []
            for index, row in df.iloc[:].iterrows():
                if row.iloc[7] == "●":
                    Hostname = row.iloc[0]
                    IPAddress = row.iloc[1]
                    Username = row.iloc[2]
                    Password = row.iloc[3]
                    EnablePass = row.iloc[4]
                    LogName = row.iloc[5]
                    if pd.isna(LogName):
                        LogName = os.path.join(log_dir, f"{Hostname}_{timestamp}.log")
                    else:
                        LogName = os.path.join(log_dir, LogName)
                    DeviceType = row.iloc[6]

                    DeviceInfo += [
                        {
                        "Hostname": Hostname,
                        "IPAddress": IPAddress,
                        "Username": Username,
                        "Password": Password,
                        "EnablePass": EnablePass,
                        "LogName": LogName,
                        "DeviceType": DeviceType
                        }
                    ]
                
                else:
                    continue
            
            return DeviceInfo
        
        except Exception as e:
            # エラーが発生した場合に警告のポップアップを表示            
            General.show_warning_popup(f"エラーが発生しました: {str(e)}")

    @keyword
    def load_testcase_from_csv(self, file_path, target_host):        

        try:
            df = pd.read_csv(file_path, encoding="utf-8")               

            TestCaseInfo = []
            for index, row in df.iloc[:].iterrows():
                check_value = row.iloc[6]
                if isinstance(check_value, str):
                    check_values = check_value.split(",")

                    if "all" in check_values or target_host in check_values:
                        ExpectedValue = row.iloc[0]
                        Operator = row.iloc[1]
                        Command = row.iloc[2]
                        TopListNum = row.iloc[3]
                        TopKey = row.iloc[4]
                        SecListNum = row.iloc[5]

                        TestCaseInfo += [   
                            {
                            "ExpectedValue": ExpectedValue,
                            "Operator": Operator,
                            "Command": Command,
                            "TopListNum": TopListNum,
                            "TopKey": TopKey,
                            "SecListNum": SecListNum,
                            }
                        ]
                    
                    else:
                        continue                            
                        
            return TestCaseInfo
            
        except Exception as e:
            # エラーが発生した場合に警告のポップアップを表示
            General.show_warning_popup(f"エラーが発生しました: {str(e)}")
    
    @keyword
    def simple_load_device_info_from_csv(self, file_path, row_number, col_number):

        try:
            # 各列の値を取得
            df = pd.read_csv(file_path, encoding="utf-8")
            
            row_number = int(row_number)
            col_number = int(col_number)

            Hostname = df.iloc[row_number, col_number]
            IPAddress = df.iloc[row_number, col_number + 1]
            Username = df.iloc[row_number, col_number + 2]
            Password = df.iloc[row_number, col_number + 3]
            EnablePass = df.iloc[row_number, col_number + 4]
            DeviceType = df.iloc[row_number, col_number + 6]

            return {
                "Hostname": Hostname,
                "IPAddress": IPAddress,
                "Username": Username,
                "Password": Password,
                "EnablePass": EnablePass,                
                "DeviceType": DeviceType
            }
        
        except Exception as e:
            # エラーが発生した場合に警告のポップアップを表示
            General.show_warning_popup(f"エラーが発生しました: {str(e)}")
